from django.shortcuts import render

# Create your views here.
from django.http import HttpResponse


from django.views.generic.edit import CreateView
from django.urls import reverse_lazy
from .models import Upload

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class UploadView(CreateView):
    model = Upload
    fields = ['upload_file', ]
    success_url = reverse_lazy('fileupload')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['documents'] = Upload.objects.all()
        return context

def generatereport(request):
    if(request.GET.get('mybtn')):
        file=( str(request.GET.get('mytextbox')) )
    df = pd.read_excel(file)
    df["Klasse"] = df["Product Variation"].str.split("|",n=4,expand=True)[2].str.replace("Klasse:","")
    df.rename(columns={"Item Name(löschen)" : "Produktname", "Anzahl ":"Anzahl"}, inplace=True)

    pd.set_option('display.max_columns', 100)  # or 1000
    pd.set_option('display.max_rows', 100)  # or 1000
    pd.set_option('display.max_colwidth', 100)
    df = pd.DataFrame(df.values.repeat(df.Anzahl, axis=0), columns=df.columns)
    df.drop(['Anzahl','Product Variation','Bestellnotiz', 'Bestellung Gesamtsumme(löschen)'], axis=1, inplace=True)
    
    t = pd.CategoricalDtype(categories=['XS', 'S','M','L','XL','XXL','XXXL'], ordered=True)
    df['Größe']=pd.Series(df.Größe, dtype=t)
    df.sort_values(by=['Klasse','Produktname','Farbe','Größe'], inplace=True,ignore_index=True)
    #=WENN(UND(I2="Ja";J2="");D2;WENN(I2="Nein";"";WENNFEHLER(RECHTS(J2;LÄNGE(J2)-50);"")))
    df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = df.apply(lambda x: x['Input Fields'] if x['Individualisierung']=='Ja' else "", axis=1)
    #df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = np.where((~df['Input Fields'].isnull()) & (~df['Individualisierung']== 'Ja') ,df['Nachnahme (Rechnungsadresse)'],"")
    df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"].str[50:]
    df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = df.apply(lambda x: x['Nachnahme (Rechnungsadresse)'] if pd.isnull(x['Individualisierungstext(zählt nur wenn Individualisierung Ja)']) else x['Individualisierungstext(zählt nur wenn Individualisierung Ja)'], axis=1)
    df["Karton"] = (df.index / 20 + 1).astype(int)
    df.drop(['Input Fields'], axis=1, inplace=True)
    
    df.sort_values(by=['Karton', 'Klasse','Produktname','Farbe','Größe'], inplace=True,ignore_index=True)
    df['Checkbox']='☐'
    df['Unterschrift']=' '
    

    df["Anzahl"]=1
    #df2= df2.pivot_table(index=['Produktname','Größe','Farbe'], 
                # columns='Individualisierung', 
                # margins = True,
                # aggfunc='size', 
                # fill_value=0)
   
    #wb = load_workbook(filename = 'vorlage_bestellliste_shop.xltx')
    #ws = wb["Orders"]
    #bestellungen = ws.tables["Bestellungen"]
    #print(bestellungen)
    
    df.columns = df.columns.astype(str)
    html = df.to_html()
    return HttpResponse(('<head><meta charset="utf-8"></head>' + html),content_type="text/html")
    